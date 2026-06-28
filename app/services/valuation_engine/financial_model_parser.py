from __future__ import annotations

import csv
from pathlib import Path
from typing import Any


XLS_WARNING = "当前建议使用 XLSX 格式，旧版 XLS 可能无法完整解析。"

SECTION_KEYWORDS: dict[str, list[str]] = {
    "收入预测表": ["收入", "Revenue", "销售", "Sales", "预测", "Forecast"],
    "成本结构表": ["成本", "Cost", "毛利", "Gross Margin", "单位经济模型"],
    "利润表": ["利润表", "P&L", "Income Statement", "损益", "净利润", "EBITDA"],
    "现金流表": ["现金流", "Cash Flow", "经营现金流", "自由现金流"],
    "资产负债表": ["资产负债表", "Balance Sheet"],
    "CAPEX / 投资计划表": ["CAPEX", "资本开支", "投资计划", "项目总投资"],
    "产能规划表": ["产能", "Capacity", "产线", "利用率"],
    "融资测算表": ["融资", "Financing", "股权结构", "融资测算", "投前", "投后"],
    "IRR / 回收期测算表": ["IRR", "回收期", "Payback", "项目测算", "NPV"],
    "敏感性分析表": ["Sensitivity", "敏感性", "情景", "保守", "中性", "乐观"],
}

FIELD_KEYWORDS: dict[str, list[str]] = {
    "历史收入": ["历史收入", "营业收入", "收入"],
    "预测收入": ["预测收入", "预计收入", "Revenue Forecast", "Forecast Revenue"],
    "收入增长率": ["收入增长率", "增长率", "Revenue Growth"],
    "产品收入拆分": ["产品收入", "产品拆分"],
    "客户收入拆分": ["客户收入", "客户拆分"],
    "单价": ["单价", "ASP", "Unit Price"],
    "销量": ["销量", "Volume", "销售量"],
    "毛利": ["毛利", "Gross Profit"],
    "毛利率": ["毛利率", "Gross Margin"],
    "净利润": ["净利润", "Net Income", "Net Profit"],
    "净利率": ["净利率", "Net Margin"],
    "EBITDA": ["EBITDA"],
    "EBITDA Margin": ["EBITDA Margin", "EBITDA率"],
    "EBIT": ["EBIT"],
    "原材料成本": ["原材料成本", "材料成本"],
    "人工成本": ["人工成本", "人员成本", "薪酬"],
    "能耗成本": ["能耗", "电费", "能源成本"],
    "折旧": ["折旧", "Depreciation"],
    "销售费用": ["销售费用", "Selling Expense", "S&M"],
    "管理费用": ["管理费用", "G&A"],
    "研发费用": ["研发费用", "R&D"],
    "财务费用": ["财务费用", "Finance Cost"],
    "OPEX": ["OPEX", "运营费用", "运营成本"],
    "经营现金流": ["经营现金流", "Operating Cash Flow"],
    "自由现金流": ["自由现金流", "Free Cash Flow", "FCF"],
    "项目现金流": ["项目现金流", "Project Cash Flow"],
    "累计现金流": ["累计现金流", "Cumulative Cash Flow"],
    "CAPEX": ["CAPEX", "资本开支"],
    "项目总投资": ["项目总投资", "总投资"],
    "建设周期": ["建设周期", "建设期"],
    "设计产能": ["设计产能", "规划产能"],
    "当前产能": ["当前产能", "现有产能"],
    "产能利用率": ["产能利用率", "利用率"],
    "产能爬坡": ["产能爬坡", "Ramp"],
    "融资金额": ["融资金额", "拟融资"],
    "投前估值": ["投前估值", "Pre-money"],
    "投后估值": ["投后估值", "Post-money"],
    "出让股权比例": ["出让股权", "股权比例"],
    "IRR": ["IRR"],
    "回收期": ["回收期", "Payback"],
    "NPV": ["NPV", "净现值"],
    "保守情景": ["保守情景", "Downside", "Conservative"],
    "中性情景": ["中性情景", "Base Case"],
    "乐观情景": ["乐观情景", "Upside", "Optimistic"],
    "单价敏感性": ["单价敏感性", "价格敏感性"],
    "销量敏感性": ["销量敏感性"],
    "毛利率敏感性": ["毛利率敏感性"],
    "利用率敏感性": ["利用率敏感性"],
    "折现率敏感性": ["折现率敏感性", "Discount Rate"],
}


def parse_financial_model(file_path: str | Path) -> dict[str, Any]:
    path = Path(file_path).expanduser()
    suffix = path.suffix.lower()
    try:
        if suffix == ".xlsx":
            return parse_xlsx(path)
        if suffix == ".csv":
            return parse_csv(path)
        if suffix == ".xls":
            return parse_xls_with_warning(path)
        return base_result(path, suffix.lstrip(".") or "unknown", "unsupported", [f"暂不支持 {suffix or '未知'} 文件类型。"], "failed")
    except Exception as exc:  # pragma: no cover - Streamlit should never crash on parser errors
        return base_result(path, suffix.lstrip(".") or "unknown", "failed", [f"财务模型解析失败：{exc}"], "failed")


def parse_xlsx(path: Path) -> dict[str, Any]:
    try:
        from openpyxl import load_workbook  # type: ignore
    except ImportError:
        return base_result(path, "xlsx", "openpyxl", ["未安装 openpyxl，无法解析 XLSX。"], "failed")

    workbook = load_workbook(path, data_only=True, read_only=True)
    sheets: list[dict[str, Any]] = []
    tables: list[dict[str, Any]] = []
    raw_preview: dict[str, list[list[Any]]] = {}
    scan_rows_by_sheet: dict[str, list[list[Any]]] = {}
    for worksheet in workbook.worksheets:
        preview_rows: list[list[Any]] = []
        scan_rows: list[list[Any]] = []
        for row_index, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
            cleaned = [clean_cell(value) for value in row]
            if any(cell != "" for cell in cleaned):
                if len(preview_rows) < 20:
                    preview_rows.append(cleaned)
                if len(scan_rows) < 200:
                    scan_rows.append(cleaned)
            if row_index >= 300:
                break
        raw_preview[worksheet.title] = preview_rows
        scan_rows_by_sheet[worksheet.title] = scan_rows
        sheets.append({"sheet_name": worksheet.title, "max_row": worksheet.max_row, "max_column": worksheet.max_column})
        tables.append({"sheet_name": worksheet.title, "rows": preview_rows})
    workbook.close()
    return build_result(path, "xlsx", "openpyxl", sheets, tables, raw_preview, scan_rows_by_sheet, [])


def parse_csv(path: Path) -> dict[str, Any]:
    warnings: list[str] = []
    rows: list[list[Any]] = []
    for encoding in ["utf-8-sig", "utf-8", "gb18030"]:
        try:
            with path.open("r", encoding=encoding, newline="") as handle:
                rows = [[clean_cell(cell) for cell in row] for row in csv.reader(handle)]
            break
        except UnicodeDecodeError:
            continue
    if not rows:
        warnings.append("CSV 文件为空或编码无法识别。")
    preview = [row for row in rows if any(cell != "" for cell in row)][:20]
    scan_rows = [row for row in rows if any(cell != "" for cell in row)][:300]
    sheet_name = "CSV"
    sheets = [{"sheet_name": sheet_name, "max_row": len(rows), "max_column": max((len(row) for row in rows), default=0)}]
    tables = [{"sheet_name": sheet_name, "rows": preview}]
    return build_result(path, "csv", "csv", sheets, tables, {sheet_name: preview}, {sheet_name: scan_rows}, warnings)


def parse_xls_with_warning(path: Path) -> dict[str, Any]:
    return base_result(path, "xls", "legacy_warning", [XLS_WARNING], "failed")


def build_result(
    path: Path,
    file_type: str,
    parser: str,
    sheets: list[dict[str, Any]],
    tables: list[dict[str, Any]],
    raw_preview: dict[str, list[list[Any]]],
    scan_rows_by_sheet: dict[str, list[list[Any]]],
    warnings: list[str],
) -> dict[str, Any]:
    detected_sections = detect_financial_sections(scan_rows_by_sheet)
    extracted_data = extract_financial_data(scan_rows_by_sheet)
    missing_data = [field for field, value in extracted_data["fields"].items() if value["extraction_result"] == "缺失"]
    quality = extraction_quality(sheets, detected_sections, extracted_data["field_assessments"])
    valuation_support = supported_models(extracted_data["fields"])
    return {
        "file_name": path.name,
        "file_path": str(path),
        "file_type": file_type,
        "parser": parser,
        "sheets": sheets,
        "tables": tables,
        "raw_preview": raw_preview,
        "detected_financial_sections": detected_sections,
        "extracted_financial_data": {
            **extracted_data,
            "missing_financial_data": missing_data,
            "requires_user_confirmation": [
                row["field"] for row in extracted_data["field_assessments"] if row["needs_confirmation"] == "是"
            ],
            "usable_financial_data": [
                field for field, value in extracted_data["fields"].items() if value["extraction_result"] != "缺失"
            ],
            "supported_valuation_models": valuation_support,
            "recommended_supplemental_materials": supplemental_materials(missing_data),
        },
        "warnings": dedupe(warnings),
        "extraction_quality": quality,
    }


def detect_financial_sections(scan_rows_by_sheet: dict[str, list[list[Any]]]) -> dict[str, list[dict[str, str]]]:
    detected: dict[str, list[dict[str, str]]] = {}
    for section, keywords in SECTION_KEYWORDS.items():
        matches = []
        for sheet_name, rows in scan_rows_by_sheet.items():
            haystack = " ".join([sheet_name, *[" ".join(str(cell) for cell in row) for row in rows[:30]]])
            matched = [keyword for keyword in keywords if keyword.lower() in haystack.lower()]
            if matched:
                matches.append({"sheet_name": sheet_name, "matched_keywords": "、".join(matched[:6])})
        detected[section] = matches
    return detected


def extract_financial_data(scan_rows_by_sheet: dict[str, list[list[Any]]]) -> dict[str, Any]:
    fields: dict[str, dict[str, Any]] = {}
    assessments: list[dict[str, str]] = []
    for field, keywords in FIELD_KEYWORDS.items():
        match = find_field(scan_rows_by_sheet, keywords)
        if match:
            value = {
                "extraction_result": match["value"],
                "source_sheet": match["sheet_name"],
                "source_position": match["position"],
                "confidence": "高",
                "needs_confirmation": "否",
            }
        else:
            value = {
                "extraction_result": "缺失",
                "source_sheet": "缺失",
                "source_position": "",
                "confidence": "缺失",
                "needs_confirmation": "是",
            }
        fields[field] = value
        assessments.append(
            {
                "field": field,
                "extraction_result": str(value["extraction_result"]),
                "source_sheet": str(value["source_sheet"]),
                "source_position": str(value["source_position"]),
                "source": "Excel明确披露" if match else "缺失",
                "confidence": str(value["confidence"]),
                "needs_confirmation": str(value["needs_confirmation"]),
            }
        )
    return {
        "fields": fields,
        "field_assessments": assessments,
        "revenue_related": pick_fields(fields, ["历史收入", "预测收入", "收入增长率", "产品收入拆分", "客户收入拆分", "单价", "销量"]),
        "gross_profit_and_profit": pick_fields(fields, ["毛利", "毛利率", "净利润", "净利率", "EBITDA", "EBITDA Margin", "EBIT"]),
        "costs_and_expenses": pick_fields(fields, ["原材料成本", "人工成本", "能耗成本", "折旧", "销售费用", "管理费用", "研发费用", "财务费用", "OPEX"]),
        "cash_flow": pick_fields(fields, ["经营现金流", "自由现金流", "项目现金流", "累计现金流"]),
        "investment_and_capacity": pick_fields(fields, ["CAPEX", "项目总投资", "建设周期", "设计产能", "当前产能", "产能利用率", "产能爬坡"]),
        "financing_and_returns": pick_fields(fields, ["融资金额", "投前估值", "投后估值", "出让股权比例", "IRR", "回收期", "NPV"]),
        "sensitivity_assumptions": pick_fields(fields, ["保守情景", "中性情景", "乐观情景", "单价敏感性", "销量敏感性", "毛利率敏感性", "利用率敏感性", "折现率敏感性"]),
    }


def find_field(scan_rows_by_sheet: dict[str, list[list[Any]]], keywords: list[str]) -> dict[str, str] | None:
    for sheet_name, rows in scan_rows_by_sheet.items():
        for row_index, row in enumerate(rows, start=1):
            for col_index, cell in enumerate(row, start=1):
                cell_text = str(cell)
                if not cell_text:
                    continue
                if any(keyword.lower() in cell_text.lower() for keyword in keywords):
                    result = first_meaningful_value(row, col_index)
                    if result:
                        return {
                            "sheet_name": sheet_name,
                            "position": f"R{row_index}C{col_index}",
                            "value": result,
                        }
                    return {
                        "sheet_name": sheet_name,
                        "position": f"R{row_index}C{col_index}",
                        "value": cell_text,
                    }
    return None


def first_meaningful_value(row: list[Any], label_col_index: int) -> str:
    for value in row[label_col_index:]:
        text = str(value).strip()
        if text and text.lower() not in {"none", "nan"}:
            return text
    return ""


def supported_models(fields: dict[str, dict[str, Any]]) -> list[str]:
    available = {field for field, value in fields.items() if value["extraction_result"] != "缺失"}
    models = []
    if {"预测收入", "历史收入"} & available:
        models.extend(["收入倍数", "可比上市公司法", "可比融资交易法"])
    if {"项目现金流", "自由现金流", "CAPEX"} & available:
        models.extend(["DCF", "项目现金流模型"])
    if "IRR" in available:
        models.append("IRR")
    if "回收期" in available:
        models.append("投资回收期")
    if {"设计产能", "当前产能", "产能利用率"} & available:
        models.extend(["产能价值法", "利用率敏感性分析"])
    if not models:
        models.append("需要补充财务数据后确认")
    return dedupe(models)


def supplemental_materials(missing_data: list[str]) -> list[str]:
    materials = []
    if any(field in missing_data for field in ["项目现金流", "自由现金流", "经营现金流"]):
        materials.append("请补充现金流预测表。")
    if any(field in missing_data for field in ["CAPEX", "项目总投资"]):
        materials.append("请补充 CAPEX / 投资计划表。")
    if any(field in missing_data for field in ["IRR", "回收期", "NPV"]):
        materials.append("请补充 IRR、NPV 或投资回收期测算。")
    if any(field in missing_data for field in ["保守情景", "中性情景", "乐观情景"]):
        materials.append("请补充保守 / 中性 / 乐观情景或敏感性分析。")
    return materials or ["请补充可核验的收入、成本、现金流、融资和估值假设。"]


def extraction_quality(
    sheets: list[dict[str, Any]],
    detected_sections: dict[str, list[dict[str, str]]],
    assessments: list[dict[str, str]],
) -> str:
    detected_count = sum(1 for matches in detected_sections.values() if matches)
    extracted_count = sum(1 for row in assessments if row["extraction_result"] != "缺失")
    if not sheets:
        return "failed"
    if detected_count >= 4 and extracted_count >= 10:
        return "high"
    if detected_count >= 2 or extracted_count >= 5:
        return "medium"
    return "low"


def pick_fields(fields: dict[str, dict[str, Any]], names: list[str]) -> dict[str, dict[str, Any]]:
    return {name: fields[name] for name in names if name in fields}


def base_result(
    path: Path,
    file_type: str,
    parser: str,
    warnings: list[str],
    extraction_quality: str,
) -> dict[str, Any]:
    return {
        "file_name": path.name,
        "file_path": str(path),
        "file_type": file_type,
        "parser": parser,
        "sheets": [],
        "tables": [],
        "raw_preview": {},
        "detected_financial_sections": {},
        "extracted_financial_data": {
            "fields": {},
            "field_assessments": [],
            "missing_financial_data": [],
            "requires_user_confirmation": [],
            "usable_financial_data": [],
            "supported_valuation_models": [],
            "recommended_supplemental_materials": [],
        },
        "warnings": dedupe(warnings),
        "extraction_quality": extraction_quality,
    }


def clean_cell(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, (int, float, bool)):
        return value
    return str(value).strip()


def dedupe(items: list[str]) -> list[str]:
    seen: set[str] = set()
    result: list[str] = []
    for item in items:
        if item and item not in seen:
            seen.add(item)
            result.append(item)
    return result
